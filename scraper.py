#!/usr/bin/env python3
"""
Generador de secuencia de códigos - The Bassment Club
Para GitHub Actions - VERSIÓN BLINDADA
"""

import asyncio
import sys
import os
from datetime import datetime
from playwright.async_api import async_playwright
from openpyxl import Workbook


CONFIG = {
    "url_login": "https://remote-locking.com/admin/login",
    "url_locks": "https://remote-locking.com/admin/locks/ras",
    "email": os.environ.get("LOCKER_EMAIL", "Nacho@thebassementclub.com"),
    "password": os.environ.get("LOCKER_PASSWORD", "FundaVerde"),
    "max_codes": int(os.environ.get("MAX_CODES", "200")),
}


def save_excel(data: dict, filename: str):
    """Guarda todos los lockers en formato horizontal."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Códigos"
    
    max_codes = max(len(codes) for codes in data.values()) if data else 0
    
    ws.cell(row=1, column=1, value="Locker #")
    ws.cell(row=1, column=2, value="Serial")
    for i in range(max_codes):
        ws.cell(row=1, column=3 + i, value=f"código {i}")
    
    for row_num, (locker_num, codes) in enumerate(sorted(data.items()), 2):
        ws.cell(row=row_num, column=1, value=locker_num)
        ws.cell(row=row_num, column=2, value=f"{locker_num:06d}")
        for i, code in enumerate(codes):
            ws.cell(row=row_num, column=3 + i, value=code)
    
    wb.save(filename)


async def click_confirm_modal(page, timeout=10000):
    """Espera y clickea el modal de confirmación de forma segura."""
    try:
        # Esperar a que el modal aparezca
        await page.wait_for_selector('.swal2-confirm', state='visible', timeout=timeout)
        
        # Esperar a que el botón esté habilitado (no disabled)
        for _ in range(50):  # Máximo 5 segundos esperando que se habilite
            btn = page.locator('.swal2-confirm')
            is_disabled = await btn.get_attribute('disabled')
            if is_disabled is None:  # No tiene atributo disabled = está habilitado
                await btn.click()
                await page.wait_for_timeout(500)
                return True
            await page.wait_for_timeout(100)
        
        # Si después de 5 segundos sigue disabled, intentar click de todas formas
        await page.locator('.swal2-confirm').click(force=True)
        await page.wait_for_timeout(500)
        return True
        
    except Exception as e:
        print(f" [modal error: {str(e)[:50]}]", end="")
        return False


async def process_locker(page, locker: int, max_codes: int) -> list:
    """Procesa un locker y devuelve sus códigos."""
    serial = f"{locker:06d}"
    codes = []
    
    try:
        # Buscar el locker
        row = page.locator(f'tr:has(td:text-is("{serial}"))')
        if await row.count() == 0:
            print(f"⚠️ No encontrado", end="")
            return []
        
        # Click en editar
        await row.locator('span.btn-success[data-action="edit"]').click()
        await page.wait_for_timeout(2000)
        
        # Ir a pestaña Codes
        await page.click('#code-tab')
        await page.wait_for_timeout(1000)
        
        # Reset a código inicial
        await page.click('button.btn-danger:has-text("Reset to Initial Code")')
        await page.wait_for_timeout(500)
        await click_confirm_modal(page)
        await page.wait_for_timeout(1000)
        
        # Obtener código inicial
        codes_section = page.locator('#codes')
        current_input = codes_section.locator('.form-group').filter(has_text="Current Code").locator('input')
        
        # Esperar a que el input tenga valor
        for _ in range(20):
            initial_code = await current_input.input_value()
            if initial_code and len(initial_code) >= 4:
                break
            await page.wait_for_timeout(200)
        
        if not initial_code:
            print("⚠️ Sin código inicial", end="")
            return []
            
        codes.append(initial_code)
        
        # Generar resto de códigos
        for i in range(1, max_codes):
            # Click en Activate Next Code
            await page.click('button.btn-danger:has-text("Activate Next Code")')
            await page.wait_for_timeout(500)
            
            # Confirmar modal
            await click_confirm_modal(page)
            await page.wait_for_timeout(800)
            
            # Leer nuevo código
            new_code = await current_input.input_value()
            
            # Verificar que cambió
            if new_code and new_code != codes[-1]:
                codes.append(new_code)
            else:
                # Si no cambió, esperar más y reintentar
                await page.wait_for_timeout(1000)
                new_code = await current_input.input_value()
                if new_code:
                    codes.append(new_code)
            
            # Detectar ciclo
            if new_code == initial_code and i > 10:
                print(f"🔁 Ciclo en {i}", end=" ")
                break
        
        # Volver a la lista
        await page.goto(CONFIG["url_locks"])
        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(1000)
        
        # Seleccionar "All" de nuevo
        await page.select_option('select[name="DataTables_Table_0_length"]', value="-1")
        await page.wait_for_timeout(2500)
            
    except Exception as e:
        print(f"❌ Error: {str(e)[:80]}", end=" ")
        # Recuperar: volver a la lista
        try:
            await page.goto(CONFIG["url_locks"])
            await page.wait_for_load_state("networkidle")
            await page.select_option('select[name="DataTables_Table_0_length"]', value="-1")
            await page.wait_for_timeout(2500)
        except:
            pass
    
    return codes


async def main():
    # Parsear argumentos
    if len(sys.argv) >= 3:
        locker_start = int(sys.argv[1])
        locker_end = int(sys.argv[2])
    else:
        locker_start = 1
        locker_end = 128
    
    lockers = list(range(locker_start, locker_end + 1))
    max_codes = CONFIG["max_codes"]
    output_file = f"codigos_lockers_{locker_start}-{locker_end}.xlsx"
    
    print("=" * 60)
    print("🔐 GENERADOR DE CÓDIGOS - VERSIÓN BLINDADA")
    print("=" * 60)
    print(f"📦 Lockers: {locker_start} - {locker_end} ({len(lockers)} total)")
    print(f"🎯 Códigos por locker: {max_codes}")
    print(f"⏱️  Inicio: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    all_data = {}
    start_time = datetime.now()
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page(viewport={"width": 1280, "height": 800})
        
        # Login
        print("\n🔐 Login...", end=" ", flush=True)
        await page.goto(CONFIG["url_login"])
        await page.wait_for_load_state("networkidle")
        await page.fill('#username', CONFIG["email"])
        await page.fill('#password', CONFIG["password"])
        await page.click('button[type="submit"]')
        await page.wait_for_timeout(5000)
        
        if "login" in page.url.lower():
            print("❌ LOGIN FALLÓ - Verificar credenciales")
            await browser.close()
            return
        print("✅")
        
        # Ir a RAS Locks
        print("📍 Navegando a RAS Locks...", end=" ", flush=True)
        await page.goto(CONFIG["url_locks"])
        await page.wait_for_load_state("networkidle")
        await page.select_option('select[name="DataTables_Table_0_length"]', value="-1")
        await page.wait_for_timeout(3000)
        print("✅")
        
        # Procesar cada locker
        for i, locker in enumerate(lockers, 1):
            print(f"\n[{i}/{len(lockers)}] 🔒 Locker {locker}...", end=" ", flush=True)
            
            codes = await process_locker(page, locker, max_codes)
            
            if codes:
                all_data[locker] = codes
                print(f"✅ {len(codes)} códigos")
            else:
                print("⚠️ Sin códigos")
            
            # Guardar progreso cada 2 lockers
            if i % 2 == 0:
                save_excel(all_data, output_file)
        
        await browser.close()
    
    # Guardar final
    save_excel(all_data, output_file)
    
    elapsed = (datetime.now() - start_time).seconds
    print(f"\n{'=' * 60}")
    print(f"✅ COMPLETADO en {elapsed//60}m {elapsed%60}s")
    print(f"📁 Archivo: {output_file}")
    print(f"📊 Lockers procesados: {len(all_data)}/{len(lockers)}")
    total_codes = sum(len(c) for c in all_data.values())
    print(f"📊 Total códigos: {total_codes}")
    print("=" * 60)


if __name__ == "__main__":
    asyncio.run(main())
