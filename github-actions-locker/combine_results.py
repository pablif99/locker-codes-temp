#!/usr/bin/env python3
"""
Combina los archivos Excel de todos los batches en uno solo.
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def combine_excels():
    print("🔄 Combinando archivos Excel...")
    
    # Buscar todos los archivos Excel en results/
    all_data = {}
    max_codes = 0
    
    for root, dirs, files in os.walk("results"):
        for file in files:
            if file.endswith(".xlsx"):
                filepath = os.path.join(root, file)
                print(f"   📂 Leyendo: {filepath}")
                
                wb = load_workbook(filepath)
                ws = wb.active
                
                # Leer filas (saltar header)
                for row in range(2, ws.max_row + 1):
                    locker_num = ws.cell(row=row, column=1).value
                    if not locker_num:
                        continue
                    
                    codes = []
                    for col in range(3, ws.max_column + 1):
                        val = ws.cell(row=row, column=col).value
                        if val:
                            codes.append(str(val))
                        else:
                            break
                    
                    if codes:
                        all_data[locker_num] = codes
                        max_codes = max(max_codes, len(codes))
                
                wb.close()
    
    print(f"\n📊 Total lockers: {len(all_data)}")
    print(f"📊 Max códigos: {max_codes}")
    
    # Crear archivo combinado
    wb = Workbook()
    ws = wb.active
    ws.title = "Códigos Taquillas"
    
    # Headers
    ws.cell(row=1, column=1, value="Locker #")
    ws.cell(row=1, column=2, value="Serial")
    for i in range(max_codes):
        ws.cell(row=1, column=3 + i, value=f"código {i}")
    
    # Data ordenada por número de locker
    for row_num, locker_num in enumerate(sorted(all_data.keys()), 2):
        codes = all_data[locker_num]
        ws.cell(row=row_num, column=1, value=locker_num)
        ws.cell(row=row_num, column=2, value=f"{locker_num:06d}")
        for i, code in enumerate(codes):
            ws.cell(row=row_num, column=3 + i, value=code)
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    for col in range(3, 3 + min(max_codes, 50)):
        ws.column_dimensions[get_column_letter(col)].width = 10
    
    output = "codigos_taquillas_COMPLETO.xlsx"
    wb.save(output)
    print(f"\n✅ Guardado: {output}")
    print(f"📁 {len(all_data)} lockers × hasta {max_codes} códigos")


if __name__ == "__main__":
    combine_excels()
