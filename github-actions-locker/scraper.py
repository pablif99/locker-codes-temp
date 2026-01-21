#!/usr/bin/env python3
"""
Generador de secuencia de códigos - The Bassment Club
Para GitHub Actions - acepta rango de lockers como parámetro
"""

import asyncio
import sys
import os
from datetime import datetime
from playwright.async_api import async_playwright
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ============== CONFIGURACIÓN ==============
CONFIG = {
    "url_login": "https://remote-locking.com/admin/login",
    "url_locks": "https://remote-locking.com/admin/locks/ras",
    "email": os.environ.get("LOCKER_EMAIL", "Nacho@thebassementclub.com"),
    "password": os.environ.get("LOCKER_PASSWORD", "FundaVerde"),
    "max_codes": int(os.environ.get("MAX_CODES", "500")),
}


def save_excel(data: dict, filename: str):
    """Guarda todos los lockers en formato horizontal."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Códigos"
    
    # Encontrar max códigos
    max_codes = max(len(codes) for codes in data.values()) if data else 0
    
    # Headers
    ws.cell(row=1, column=1, value="Locker #")
    ws.cell(row=1, column=2, value="Serial")
    for i in range(max_codes):
        ws.cell(row=1, column=3 + i, value=f"código {i}")
    
    # Data
    for row_num, (locker_num, codes) in enumerate(sorted(data.items()), 2):
        ws.cell(row=row_num, column=1, value=locker_num)
        ws.cell(row=row_num, column=2, value=f"{locker_num:06d}")
        for i, code in enumerate(codes):
            ws.cell(row=row_num, column=3 + i, value=code)
    
    wb.save(filename)


async def process_locker(page, locker: int, max_codes: int) -> list:
    """Procesa un locker y devuelve sus códigos."""
    serial = f"{locker:06d}"
    codes = []
    
    try:
        # Buscar y clickear el locker
        row = page.locator(f'tr:has(td:text-is("{serial}"))')
        if await row.count() == 0:
            print(f"   ⚠️ Locker {locker} no encontrado")
            return []
        
        await row.locator('span.btn-success[data-action="edit"]').click()
        await page.wait_for_timeout(1500)
        
        # Ir a pestaña Codes
        await page.click('#code-tab')
        await page.wait_for_timeout(500)
        
        # Reset
        await page.click('button.btn-danger:has-text("Reset to Initial Code")')
        await page.wait_for_timeout(800)
        confirm = page.locator('.swal2-confirm')
        if await confirm.count() > 0:
            await confirm.click()
            await page.wait_for_timeout(800)
        
        # Obtener código inicial
        codes_section = page.locator('#codes')
        current_input = codes_section.locator('.form-group').filter(has_text="Current Code").locator('input')
        await page.wait_for_timeout(500)
        initial_code = await current_input.input_value()
        codes.append(initial_code)
        
        # Generar resto de códigos
        for i in range(1, max_codes):
            await page.click('button.btn-danger:has-text("Activate Next Code")')
            await page.wait_for_timeout(800)
            
            confirm = page.locator('.swal2-confirm')
            if await confirm.count() > 0:
                await confirm.click()
                await page.wait_for_timeout(800)
            
            new_code = await current_input.input_value()
            codes.append(new_code)
            
            # Detectar ciclo (código vuelve al inicial)
            if new_code == initial_code and i > 10:
                print(f"   🔁 Ciclo detectado en código {i}")
                break
        
        # Volver a la lista
        back_btn = page.locator('a.btn:has-text("Back")')
        if await back_btn.count() > 0:
            await back_btn.click()
        else:
            await page.goto(CONFIG["url_locks"])
        
        await page.wait_for_load_state("networkidle")
        await page.wait_for_timeout(1000)
        
        # Re-seleccionar All
        entries_select = page.locator('select[name="DataTables_Table_0_length"]')
        current_value = await entries_select.input_value()
        if current_value != "-1":
            await entries_select.select_option(value="-1")
            await page.wait_for_timeout(2000)
            
    except Exception as e:
        print(f"   ❌ Error: {e}")
        # Intentar volver a la lista
        try:
            await page.goto(CONFIG["url_locks"])
            await page.wait_for_load_state("networkidle")
            await page.select_option('select[name="DataTables_Table_0_length"]', value="-1")
            await page.wait_for_timeout(2000)
        except:
            pass
    
    return codes


async def main():
    # Parsear argumentos: locker_start locker_end
    if len(sys.argv) >= 3:
        locker_start = int(sys.argv[1])
        locker_end = int(sys.argv[2])
    else:
        locker_start = 1
        locker_end = 128
    
    lockers = list(range(locker_start, locker_end + 1))
    max_codes = CONFIG["max_codes"]
    output_file = f"codigos_lockers_{locker_start}-{locker_end}.xlsx"
    
    print("=" * 60)
    print("🔐 GENERADOR DE CÓDIGOS - GitHub Actions")
    print("=" * 60)
    print(f"📦 Lockers: {locker_start} - {locker_end} ({len(lockers)} total)")
    print(f"🎯 Códigos por locker: {max_codes}")
    print(f"⏱️  Inicio: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    all_data = {}
    start_time = datetime.now()
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page(viewport={"width": 1280, "height": 800})
        
        # Login
        print("\n🔐 Login...", end=" ", flush=True)
        await page.goto(CONFIG["url_login"])
        await page.wait_for_load_state("networkidle")
        await page.fill('#username', CONFIG["email"])
        await page.fill('#password', CONFIG["password"])
        await page.click('button[type="submit"]')
        await page.wait_for_timeout(3000)
        
        if "login" in page.url.lower():
            print("❌ FALLÓ")
            await browser.close()
            return
        print("✅")
        
        # Ir a RAS Locks
        print("📍 Navegando a RAS Locks...", end=" ", flush=True)
        await page.goto(CONFIG["url_locks"])
        await page.wait_for_load_state("networkidle")
        await page.select_option('select[name="DataTables_Table_0_length"]', value="-1")
        await page.wait_for_timeout(3000)
        print("✅")
        
        # Procesar cada locker
        for i, locker in enumerate(lockers, 1):
            elapsed = (datetime.now() - start_time).seconds
            print(f"\n[{i}/{len(lockers)}] 🔒 Locker {locker}...", end=" ", flush=True)
            
            codes = await process_locker(page, locker, max_codes)
            
            if codes:
                all_data[locker] = codes
                print(f"✅ {len(codes)} códigos")
            else:
                print("⚠️ Sin códigos")
            
            # Guardar progreso cada 5 lockers
            if i % 5 == 0:
                save_excel(all_data, output_file)
                print(f"   💾 Progreso guardado ({i}/{len(lockers)})")
        
        await browser.close()
    
    # Guardar final
    save_excel(all_data, output_file)
    
    elapsed = (datetime.now() - start_time).seconds
    print(f"\n{'=' * 60}")
    print(f"✅ COMPLETADO en {elapsed//60}m {elapsed%60}s")
    print(f"📁 Archivo: {output_file}")
    print(f"📊 Lockers procesados: {len(all_data)}/{len(lockers)}")
    print("=" * 60)


if __name__ == "__main__":
    asyncio.run(main())
